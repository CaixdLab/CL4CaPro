import os
import openpyxl
import csv

ClincFilePath = 'TCGA-CDR-SupplementalTableS1.xlsx'

if __name__ == "__main__":

    #Check Path
    if os.path.exists(ClincFilePath):
        clinic_workbook = openpyxl.load_workbook(ClincFilePath)
        EEP_worksheet = clinic_workbook['ExtraEndpoints']
    else:
        print('Invalid Path for the clinic data')

    #Get Data
    #CancerType = ['BLCA', 'CESC', 'ESCA', 'HNSC', 'LUSC', 'LUAD']
    #CancerType = ['BRCA', 'KIRC', 'LGG', 'LIHC', 'OV', 'STAD'] #Cox-nnet
    CancerType = ['LAML', 'ACC', 'CHOL', 'LCML', 'COAD', 'CNTL', 'FPPP', 'GBM', 'KICH', 'KIRP', 'DLBC', 'MESO', 'MISC', 'PAAD', 'PCPG', 'PRAD', 'READ', 'SARC', 'SKCM', 'STAD', 'TGCT', 'THYM', 'THCA', 'UCS', 'UCEC', 'UVM'] #TCGAs

    for eachCancer in CancerType:
        ExtractData = []
        RowInfo = []
        for col in range(0, EEP_worksheet.max_column):
            RowInfo.append(EEP_worksheet[chr(col + ord('A')) + '1'].value)
        ExtractData.append(RowInfo)
        for row in range(2, EEP_worksheet.max_row + 1):
            if EEP_worksheet['C' + str(row)].value == eachCancer and (EEP_worksheet['K' + str(row)].value != '#N/A' and int(EEP_worksheet['K' + str(row)].value) != 0): #E for PFI.time.1
                RowInfo = []
                for col in range(0, EEP_worksheet.max_column):
                    RowInfo.append(EEP_worksheet[chr(col + ord('A')) + str(row)].value)
                ExtractData.append(RowInfo)

        #Save CSV
        with open(eachCancer + 'Data.csv', 'w', newline="") as f:
            writer = csv.writer(f)
            for item in ExtractData:
                writer.writerow(item)
        print('Cancer: ', eachCancer, ' Len: ', len(ExtractData) - 1)



