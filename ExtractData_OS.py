import os
import openpyxl
import csv

OSFilePath = 'Clean_mmc.xlsx'

if __name__ == "__main__":

    #Check Path
    if os.path.exists(OSFilePath):
        os_workbook = openpyxl.load_workbook(OSFilePath)
        EEP_worksheet = os_workbook['PanImmune_MS']
    else:
        print('Invalid Path for the OS data')

    #Get Data
    CancerType = ['LGG']
    for eachCancer in CancerType:
        ExtractData = []
        RowInfo = []
        for col in range(0, EEP_worksheet.max_column):
            RowInfo.append(EEP_worksheet[chr(col + ord('A')) + '1'].value)
        ExtractData.append(RowInfo)
        for row in range(2, EEP_worksheet.max_row + 1):
            if EEP_worksheet['B' + str(row)].value == eachCancer and (EEP_worksheet['J' + str(row)].value != 'NA' and int(EEP_worksheet['J' + str(row)].value) != 0): #E for PFI.time.1
                RowInfo = []
                for col in range(0, EEP_worksheet.max_column):
                    RowInfo.append(EEP_worksheet[chr(col + ord('A')) + str(row)].value)
                ExtractData.append(RowInfo)

        #Save CSV
        with open('OS_' + eachCancer + 'Data.csv', 'w', newline="") as f:
            writer = csv.writer(f)
            for item in ExtractData:
                writer.writerow(item)
        print('Cancer: ', eachCancer, ' Len: ', len(ExtractData) - 1)



